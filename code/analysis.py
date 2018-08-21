import matplotlib.pyplot as plt, numpy as np, matplotlib, datetime as dtm
import re,urllib2,os
from scipy.optimize import curve_fit
from scipy.interpolate import interp1d
from scipy.special import erfc
import sqlite3
from npat import plotter
_mcnp_path = '/home/jmorrell/MCNP'
_db_connection = sqlite3.connect('../data/peak_data.db')
_db = _db_connection.cursor()


class isotope(object):
	def __init__(self,istp):
		self.db_connection = _db_connection
		self.db = _db
		self.tex_g,self.metastable = False,False
		if istp.endswith('m'):
			self.metastable = True
			self.istp = istp[:-1].upper()
		elif istp.endswith('g'):
			self.istp = istp[:-1].upper()
			self.tex_g = True
		else:
			self.istp = istp.upper()
		self.t_half,self.unc_t_half = None,None
	def parse_isotope_dat(self,ln):
		fmt = {}
		for n,el in enumerate(['isotope','E_parent','t_half','energy','intensity','unc_intensity','unc_t_half']):
			fmt[el] = str(ln[n]) if n==0 else float(ln[n])
		fmt['metastable'] = True if fmt['E_parent'] > 0 else False
		return fmt
	def name(self):
		return self.istp+('m' if self.metastable else '')+('g' if self.tex_g else '')
	def TeX(self):
		el = ''.join(re.findall('[A-Z]+',self.istp))
		return r'$^{'+self.istp.split(el)[0]+('m' if self.metastable else '')+('g' if self.tex_g else '')+r'}$'+el.title()
	def gammas(self,Imin=0,Emin=70):
		self.gm = []
		for ln in self.db.execute('SELECT * FROM ensdf WHERE isotope=?',(self.istp,)):
			fmt = self.parse_isotope_dat(ln)
			if fmt['metastable']==self.metastable:
				### NOTE: WILL BE WRONG IF MORE THAN ONE ISOMER!!! ###
				self.t_half,self.unc_t_half = fmt['t_half'],fmt['unc_t_half']
				self.gm.append((fmt['energy'],fmt['intensity'],fmt['unc_intensity']))
		return [g for g in self.gm if g[0]>=Emin and g[1]>=Imin]
	def half_life(self,units='s'):
		if self.t_half is None:
			self.t_half = [self.parse_isotope_dat(ln) for ln in self.db.execute('SELECT * FROM ensdf WHERE isotope=? AND E_parent'+('>0' if self.metastable else '=0'),(self.istp,))][0]['t_half']
		return self.t_half/{'s':1.0,'m':60.0,'h':3600.0,'d':86400.0,'y':3.154e+7}[units]
	def unc_half_life(self,units='s'):
		if self.unc_t_half is None:
			self.unc_t_half = [self.parse_isotope_dat(ln) for ln in self.db.execute('SELECT * FROM ensdf WHERE isotope=? AND E_parent'+('>0' if self.metastable else '=0'),(self.istp,))][0]['unc_t_half']
		return self.unc_t_half/{'s':1.0,'m':60.0,'h':3600.0,'d':86400.0,'y':3.154e+7}[units]
	def decay_const(self,units='s'):
		return np.log(2)/self.half_life(units)
	def unc_decay_const(self,units='s'):
		return np.log(2)*self.unc_half_life(units)/self.half_life(units)**2
	def is_metastable(self):
		return self.metastable


class spectrum(object):
	def __init__(self,filename):
		end_of_beam = dtm.datetime(2017,9,8,0,4,14)
		self.db_connection = _db_connection
		self.db = _db
		self.filename = filename.split('.')[0]
		directory = [str(i[1]) for i in self.db.execute('SELECT * FROM all_files WHERE filename=?',(filename,))][0]
		f = open(directory+filename,'r').read().split('\n')
		self.start_time = dtm.datetime.strptime(f[7].replace('\r',''),'%m/%d/%Y %H:%M:%S')
		self.live_time = float(f[9].split(' ')[0].strip())
		self.real_time = float(f[9].split(' ')[1].strip())
		self.spec = [int(f[n].strip()) for n in range(12,13+int(f[11].split(' ')[1]))]
		self.cooling_time = (self.start_time-end_of_beam).total_seconds()
		self.sample = self.filename.split('_')[2]
		self.distance = float(self.filename.split('_')[1].replace('p','.')[:-2])
		self.peak_fits = None
		self.calibration = [[float(i[1].split(',')[0]),float(i[1].split(',')[1])] for i in self.db.execute('SELECT * FROM calibration WHERE distance=?',(self.distance,))][0]
		self.efficiency = [[float(i[2].split(',')[0]),float(i[2].split(',')[1]),float(i[2].split(',')[2])] for i in self.db.execute('SELECT * FROM calibration WHERE distance=?',(self.distance,))][0]
		self.resolution = [float(i[3]) for i in self.db.execute('SELECT * FROM calibration WHERE distance=?',(self.distance,))][0]
		self.pk_args = [(float(i[4]),float(i[5])) for i in self.db.execute('SELECT * FROM calibration WHERE distance=?',(self.distance,))][0]
		self.SNP = self.SNIP()
		self.isotope_list = [str(i[1]).split(',') for i in self.db.execute('SELECT * FROM sample_map WHERE sample=?',(self.sample[:2],))][0]
		self.pallate = plotter().pallate()
	def save_as_radware(self):
		from struct import pack
		with open('rw_files/'+self.filename+'.spe','wb') as f:
			f.write(pack('I8s6I'+str(len(self.spec))+'fI',*[24,self.filename[:8],len(self.spec),1,1,1,24,len(self.spec)*4]+[float(i) for i in self.spec]+[len(self.spec)*4]))
	def get_energy(self):
		a,b = self.calibration[0],self.calibration[1]
		return [a*i+b for i in range(len(self.spec))]
	def get_plot_spec(self,logscale=True,mn=0,mx=None):
		mx = len(self.spec) if mx is None else mx
		lim_y = 1 if logscale else 0
		bn = [max((self.spec[mn],lim_y))]
		for i in range(mn,mx-1):
			bn.append(max((self.spec[i],lim_y)))
			bn.append(max((self.spec[i+1],lim_y)))
		return bn
	def get_plot_energy(self,mn=0,mx=None):
		mx = len(self.spec) if mx is None else mx
		a,b = self.calibration
		eng = [a*(i-0.5)+b for i in range(mn,mx)]
		bn = []
		for i in range(1,len(eng)):
			bn.append(eng[i-1])
			bn.append(eng[i])
		bn.append(eng[-1])
		return bn
	def get_efficiency(self,energy):
		a,b,c = self.efficiency[0],self.efficiency[1],self.efficiency[2]
		return np.exp(a*np.log(energy)**2+b*np.log(energy)+c)
	def set_calibration(self,calibration):
		self.calibration = calibration
	def set_efficiency(self,efficiency):
		self.efficiency = efficiency
	def exp_smooth(self,ls,alpha=0.3):
		R,RR,b = [ls[0]],[ls[-1]],1.0-alpha
		for i,ii in zip(ls[1:],reversed(ls[:-1])):
			R.append(alpha*i+b*R[-1])
			RR.append(alpha*ii+b*RR[-1])
		return [0.5*(R[n]+r) for n,r in enumerate(reversed(RR))]
	def SNIP(self,N=9,alpha=0.2):
		dead=0
		while self.spec[dead]==0:
			dead+=1
		vi = [np.log(np.log(np.sqrt(i+1.0)+1.0)+1.0) for i in self.spec]
		a,L,off = self.resolution,len(vi),int(self.resolution*N*len(vi)**0.5)
		for M in range(1,N):
			vi = vi[:dead+off]+[min((v,0.5*(vi[n-max((M*int(a*n**0.5),M/2))]+vi[n+max((M*int(a*n**0.5),M/2))]))) for v,n in zip(vi[dead+off:L-off],range(dead+off,L-off))]+vi[-1*off:]
		return [s+1.5*np.sqrt(s+1) for n,s in enumerate(self.exp_smooth([int((np.exp(np.exp(i)-1.0)-1.0)**2)-1 for i in vi],alpha=alpha))]
	def find_pks(self,sensitivity=0.5,ls=None):
		N,sg,alpha,Emin = int(10-4*sensitivity),4.5-4.0*sensitivity,0.1+0.5*sensitivity,50
		SNP = self.SNIP(N=N)
		ls = self.spec if ls is None else ls
		clip = [int(i-SNP[n]) if (i-SNP[n])>sg*np.sqrt(SNP[n]) else 0 for n,i in enumerate(self.exp_smooth(ls,alpha=alpha))]
		pks = [n+1 for n,i in enumerate(clip[1:-1]) if all([i>clip[n],i>clip[n+2],clip[n]>0,clip[n+2]>0])]
		return [p for n,p in enumerate(pks[1:]) if pks[n]+3*self.resolution*pks[n]**0.5<p and p*self.calibration[0]+self.calibration[1]>Emin]
	def simple_regression(self,x,y):
		xb,yb = np.average(x),np.average(y)
		m = sum([(i-xb)*(y[n]-yb) for n,i in enumerate(x)])/sum([(i-xb)**2 for i in x])
		return m,yb-m*xb
	def peak(self,x,A,mu,sig):
		r2,R,alpha = 1.41421356237,self.pk_args[0],self.pk_args[1]
		return [A*np.exp(-0.5*((i-mu)/sig)**2)+R*A*np.exp((i-mu)/(alpha*sig))*erfc((i-mu)/(r2*sig)+1.0/(r2*alpha)) for i in x]
	def Npeak(self,x,*args):
		N = (len(args)-2)/3
		I = [self.peak(x,args[2+3*n],args[2+3*n+1],args[2+3*n+2]) for n in range(N)]
		return [args[0]*i+args[1]+sum([I[m][n] for m in range(N)]) for n,i in enumerate(x)]
	def chi2(self,fn,x,b,y):
		if float(len(y)-len(b))<1:
			return float('Inf')
		return sum([(y[n]-i)**2/y[n] for n,i in enumerate(fn(x,*b)) if y[n]>0])/float(len(y)-len(b))
	def guess_A0(self):
		print 'Fitting Peaks in',self.filename
		A0,gammas = [],[]
		a,b,L = self.calibration[0],self.calibration[1],len(self.spec)
		for itp in self.isotope_list:
			ip = isotope(itp)
			gammas += [(ip.name(),ip.decay_const())+g for g in ip.gammas(Imin=0.11) if int((g[0]-b)/a)<L]
		clip = [int(i-self.SNP[n]) if int(i-self.SNP[n])>max((1.5,1.5*np.sqrt(self.SNP[n]))) else 0 for n,i in enumerate(self.exp_smooth(self.spec,alpha=0.35))]
		gammas = [g+(int((g[2]-b)/a),) for g in gammas if clip[int((g[2]-b)/a)]>0 and abs(g[2]-511.0)>10] # self.cooling_time<=15.0*np.log(2)/g[1]
		N_c = [int(2.506*self.resolution*g[-1]**0.5*clip[g[-1]])+int(2.0*self.pk_args[0]*self.pk_args[1]*self.resolution*g[-1]**0.5*clip[g[-1]]*np.exp(-0.5/self.pk_args[1]**2)) for g in gammas]
		A0 = [N_c[n]/(g[3]*self.get_efficiency(g[2])) for n,g in enumerate(gammas)]
		itps = {i:[] for i in list(set([g[0] for g in gammas]))}
		for n,g in enumerate(gammas):
			itps[g[0]].append((A0[n],g[3],N_c[n]))
		return {itp:round(np.average([i[0] for i in a],weights=[np.sqrt(i[2]*(i[0]/i[2])**2+i[1]**2) for i in a]),0) for itp,a in itps.iteritems()}
	def get_p0(self,A0=None):
		if A0 is None:
			A0 = self.guess_A0()
		if len(A0)==0:
			return []
		gammas = []
		a,b,L,RS = self.calibration[0],self.calibration[1],len(self.spec),self.resolution
		for itp,A in A0.iteritems():
			ip = isotope(itp)
			gammas += [(itp,A,ip.decay_const(),((g[0]-b)/a))+g for g in ip.gammas(Imin=0.11) if ((g[0]-b)/a)<L and abs(g[0]-511.0)>10]
		N_c = [g[1]*(self.get_efficiency(g[4])*g[5]) for g in gammas]
		A = [int(N_c[n]/(2.506*RS*g[3]**0.5+2*self.pk_args[0]*self.pk_args[1]*RS*g[3]**0.5*np.exp(0.5/self.pk_args[1]**2))) for n,g in enumerate(gammas)]
		gammas = sorted([g+(int(g[3]-6.0*RS*g[3]**0.5),int(g[3]+5.5*RS*g[3]**0.5),A[n]) for n,g in enumerate(gammas) if A[n]>max((1.5,1.5*np.sqrt(self.SNP[int(g[3])])))],key=lambda h:h[4])
		pks = [[gammas[0]]]
		for g in gammas[1:]:
			if g[7]<pks[-1][-1][8]:
				pks[-1].append(g)
			else:
				pks.append([g])
		p0 = []
		for p in pks:
			m,b = self.simple_regression(range(p[0][7],p[-1][8]),self.SNP[p[0][7]:p[-1][8]])
			bm = max((self.SNP[int(p[0][3])]**0.5,1.0))
			p0.append({'l':p[0][7],'h':p[-1][8]+1,'p0':[m,b],'pk_info':[],'bounds':([m-1e-5,b-1.5*bm],[m+1e-5,b+1.5*bm])})
			for gm in p:
				p0[-1]['pk_info'].append({'istp':gm[0],'eng':gm[4],'I':gm[5],'unc_I':gm[6],'lm':gm[2]})
				p0[-1]['p0'] += [gm[9],gm[3],RS*gm[3]**0.5]
				p0[-1]['bounds'] = (p0[-1]['bounds'][0]+[0,gm[3]-0.25*RS*gm[3]**0.5-2.0,0.75*RS*gm[3]**0.5],p0[-1]['bounds'][1]+[7.5*abs(gm[9])+1,gm[3]+0.25*RS*gm[3]**0.5+2.0,1.5*RS*gm[3]**0.5])
		return p0
	def filter_fits(self,fits,loop=True):
		good_fits = []
		RS = self.resolution
		for f in fits:
			N = (len(f['fit'])-2)/3
			chi2 = self.chi2(self.Npeak,range(f['l'],f['h']),f['fit'],self.spec[f['l']:f['h']])
			for n,p in enumerate(f['pk_info']):
				ft,u = f['fit'],f['unc']
				p['N'] = int(ft[2+3*n]*(2.506*ft[4+3*n]+2*self.pk_args[0]*self.pk_args[1]*ft[4+3*n]*np.exp(-0.5/self.pk_args[1]**2)))
				p['unc_N'] = int(np.sqrt(abs(u[2+3*n][2+3*n])*(p['N']/ft[2+3*n])**2+abs(u[4+3*n][4+3*n])*(p['N']/ft[4+3*n])**2))
			keep = [n for n in range(N) if f['fit'][2+3*n]>max((2.5,2.5*np.sqrt(self.SNP[int(f['fit'][3+3*n])]))) and (f['pk_info'][n]['unc_N']<25*f['pk_info'][n]['N']) and chi2<100.0]
			for n in keep:
				arr = [0,1,2+3*n,3+3*n,4+3*n]
				good_fits.append({'fit':[f['fit'][i] for i in arr]})
				good_fits[-1]['unc'] = [[f['unc'][i][j] for j in arr] for i in arr]
				good_fits[-1]['p0'] = [f['p0'][i] for i in arr]
				good_fits[-1]['pk_info'] = [f['pk_info'][n]]
				good_fits[-1]['bounds'] = ([f['bounds'][0][i] for i in arr],[f['bounds'][1][i] for i in arr])
				good_fits[-1]['l'],good_fits[-1]['h'] = int(f['fit'][3+3*n]-5.0*f['fit'][4+3*n]),int(f['fit'][3+3*n]+4.5*f['fit'][4+3*n])
		good_fits = sorted(good_fits,key=lambda h:h['fit'][3])
		if len(good_fits)==0:
			self.peak_fits = []
			return []
		groups = [good_fits[0]]
		for g in good_fits[1:]:
			if g['l']<groups[-1]['h']:
				L = len(groups[-1]['fit'])
				groups[-1]['fit'] += g['fit'][2:]
				groups[-1]['unc'] = [gp+[0,0,0] for gp in groups[-1]['unc']]+[[0 for i in range(L)]+gp[2:] for gp in g['unc'][2:]]
				groups[-1]['p0'] += g['p0'][2:]
				groups[-1]['pk_info'].append(g['pk_info'][0])
				groups[-1]['bounds'] = (groups[-1]['bounds'][0]+g['bounds'][0][2:],groups[-1]['bounds'][1]+g['bounds'][1][2:])
				groups[-1]['h'] = g['h']
			else:
				groups.append(g)
		for g in groups:
			chi2 = self.chi2(self.Npeak,range(g['l'],g['h']),g['fit'],self.spec[g['l']:g['h']])
			for n,p in enumerate(g['pk_info']):
				p['chi2'] = chi2
		self.peak_fits = None if loop else groups
		return self.filter_fits(groups,loop=False) if loop else groups
	def calc_A0(self,fits):
		gammas = []
		for f in fits:
			gammas += f['pk_info']
		A0 = [g['N']/(g['I']*self.get_efficiency(g['eng'])) for g in gammas]
		itps = {i:[] for i in list(set([g['istp'] for g in gammas]))}
		for n,g in enumerate(gammas):
			itps[g['istp']].append((A0[n],np.sqrt(g['unc_N']**2*(A0[n]/g['N'])**2+g['unc_I']**2*(A0[n]/g['I'])**2)))
		return {itp:np.average([i[0] for i in a],weights=[1.0/i[1]**2 for i in a]) for itp,a in itps.iteritems()}
	def fit_peaks(self,A0=None):
		if A0 is None and self.peak_fits is not None:
			return self.peak_fits
		p0 = self.get_p0(A0)
		fits = []
		for p in p0:
			try:
				p['fit'],p['unc'] = curve_fit(self.Npeak,range(p['l'],p['h']),self.spec[p['l']:p['h']],p0=p['p0'],bounds=p['bounds'],sigma=np.sqrt(self.SNP[p['l']:p['h']]))
				fits.append(p)
			except Exception, err:
				print 'Error on peak:',p['pk_info']
				print Exception, err
		if A0 is None:
			return self.fit_peaks(A0=self.calc_A0(self.filter_fits(fits)))
		return self.filter_fits(fits)
	def plot_fits(self,logscale=True,wfit=True,bg=False,save=False,zoom=False,subpeak=False,printout=False):
		f,ax = plt.subplots() if not save else plt.subplots(figsize=(12.8,4.8))
		mn,mx = 0,len(self.spec)
		if wfit:
			for n,pk in enumerate(self.fit_peaks()):
				if printout:
					print pk['pk_info']
				if zoom:
					mn,mx = pk['l']-20 if mn==0 else mn,pk['h']+20
				a,b = self.calibration[0],self.calibration[1]
				ax.plot([a*i+b for i in np.arange(pk['l'],pk['h'],0.1)],self.Npeak(np.arange(pk['l'],pk['h'],0.1),*pk['fit']),lw=1.2,color=self.pallate['r'],zorder=10,label=('Peak Fit'+('s' if len(self.peak_fits)>1 else '') if n==0 else None))
				if subpeak:
					N = (len(pk['fit'])-2)/3
					if N>1:
						for m in range(N):
							p = pk['fit'][:2]+pk['fit'][2+m*3:5+m*3]
							ax.plot([a*i+b for i in np.arange(pk['l'],pk['h'],0.1)],self.Npeak(np.arange(pk['l'],pk['h'],0.1),*p),lw=1.0,ls='--',color=self.pallate['r'],zorder=5)
		if bg:
			ax.plot(self.get_energy()[mn:mx],self.SNP[mn:mx],lw=1.0,color=self.pallate['b'],label='Background',zorder=3)
		ax.plot(self.get_plot_energy(mn,mx),self.get_plot_spec(logscale,mn,mx),lw=1.0,color=self.pallate['k'],label=self.filename,zorder=1)
		if logscale:
			ax.set_yscale('log')
		ax.set_ylim((max((0.9,ax.get_ylim()[0])),ax.get_ylim()[1]))
		ax.set_xlabel('Energy (keV)',fontsize=20)
		ax.set_ylabel('Counts',fontsize=20)
		ax.legend(loc=0)
		f.tight_layout()
		if save:
			f.savefig('../plots/peak_fits/'+self.filename+'_fits.png')
			f.savefig('../plots/peak_fits/'+self.filename+'_fits.pdf')
			plt.close()
		else:
			plt.show()



class calibration(object):
	def __init__(self):
		self.db = _db
		self.db_connection = _db_connection
		fnms = [str(i[0]) for i in self.db.execute('SELECT * FROM all_files WHERE directory LIKE "%Calib%"')]
		self.spectra = [spectrum(s) for s in fnms]
		self.pallate =plotter().pallate()
	def objective(self,m,b):
		L = len(self.log_peaks[0])
		I = [[int((e-b)/m) for e in eng] for eng in self.calib_peak_eng]
		return sum([sum([s[i] for i in I[n] if 0<i<L]) for n,s in enumerate(self.log_peaks)])
	def guess_energy_cal(self):
		# first guesses based on peak location, then CMA-ES optimization
		CS_spec = spectrum('AL170825_10cm_Cs137.Spe')
		guess = (661.657/list(sorted([(p,CS_spec.spec[p]-CS_spec.SNP[p]) for p in CS_spec.find_pks()],key=lambda h:h[1]))[-1][0],0.0)
		delta0 = (0.1*guess[0],20.0)
		gammas = [isotope(s.isotope_list[0]).gammas(Imin=0.3,Emin=70) for s in self.spectra]
		self.calib_peak_eng = [[g[0] for g in gm] for gm in gammas]
		self.log_peaks = [[np.log(np.log(np.sqrt(max((i-s.SNP[n],0))+1.0)+1.0)+1.0) for n,i in enumerate(s.spec)] for s in self.spectra]
		K = 100
		sm = sum([1.0/float(i)**2 for i in range(1,K+1)])
		wts = [1.0/(sm*float(i)**2) for i in range(1,K+1)]
		best_guess = (self.objective(guess[0],guess[1]),guess)
		for i in range(10):
			p0 = [(guess[0]+delta0[0]*g*np.exp(-i/4.0),guess[1]+delta0[1]*g*np.exp(-i/4.0)) for g in np.random.normal(0,1,K)]
			G = sorted([(self.objective(p[0],p[1]),p) for p in p0],key=lambda h:h[0],reverse=True)
			if G[0][0]>best_guess[0]:
				best_guess = G[0]
			guess = (sum([g[1][0]*wts[n] for n,g in enumerate(G)]),sum([g[1][1]*wts[n] for n,g in enumerate(G)]))
		return best_guess[1]
	def fits(self,usecal=True):
		guess = self.spectra[0].calibration if usecal else self.guess_energy_cal()
		self.db.execute('DELETE FROM calibration_peaks')
		for s in self.spectra:
			ip = isotope(s.isotope_list[0])
			s.set_calibration([guess[0],guess[1]])
			for pk in s.fit_peaks():
				N = (len(pk['p0'])-2)/3
				for n in range(N):
					w = 'pk_info'
					pkinfo = [s.filename+'.Spe',ip.istp,s.distance,pk[w][n]['N'],pk[w][n]['unc_N'],pk[w][n]['chi2'],pk[w][n]['eng'],pk[w][n]['I'],pk[w][n]['unc_I'],pk['fit'][3+n*3],np.sqrt(pk['unc'][3+n*3][3+n*3])]
					pkinfo += [pk['fit'][4+n*3],np.sqrt(pk['unc'][4+n*3][4+n*3])]
					if pk[w][n]['unc_N']/pk[w][n]['N']<0.02:
						self.db.execute('INSERT INTO calibration_peaks VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',tuple(pkinfo))
			s.plot_fits(wfit=True,logscale=True,save=True,zoom=True)
			self.db_connection.commit()
	def energy_calibration(self,show=False):
		guess = self.spectra[0].calibration
		pks = [[float(i[0]),float(i[1]),float(i[2])] for i in self.db.execute('SELECT energy,mu,unc_mu FROM calibration_peaks ORDER BY "energy" ASC') if float(i[2])<float(i[1])]
		fit,unc = curve_fit(lambda x,m,b:[m*i+b for i in x],[p[1] for p in pks],[p[0] for p in pks],p0=[guess[0],guess[1]],sigma=[guess[0]*p[2] for p in pks])
		for dist in [float(i[0]) for i in self.db.execute('SELECT * FROM calibration')]:
			self.db.execute('UPDATE calibration SET energy_calibration=? WHERE distance=?',(','.join([str(i) for i in fit]),dist))
		self.db_connection.commit()
		f,ax = plt.subplots(2,sharex=True,gridspec_kw = {'height_ratios':[3, 2]})
		xbins = np.arange(pks[0][1]-40,pks[-1][1]+20,0.5)
		ax[0].plot(xbins,[fit[0]*i+fit[1] for i in xbins],lw=2.0,color=self.pallate['gy'],label='Linear Fit')
		ax[0].errorbar([p[1] for p in pks],[p[0] for p in pks],yerr=[fit[0]*p[2] for p in pks],color=self.pallate['k'],ls='None',marker='o',label='Peak Centroids')
		ax[0].set_ylabel('Energy (keV)')
		ax[0].text(2200,150,r'$E_{\gamma}='+str(round(fit[0],3))+r'\cdot N'+('+' if fit[1]>0 else '')+str(round(fit[1],3))+r'$')
		ax[0].legend(loc=0)
		ax[1].plot(xbins,[0.0 for i in xbins],ls='--',color=self.pallate['gy'],lw=2.0)
		ax[1].errorbar([p[1] for p in pks],[p[0]-fit[0]*p[1]-fit[1] for p in pks],yerr=[fit[0]*p[2] for p in pks],color=self.pallate['k'],ls='None',marker='o')
		label = r'$\sigma_E='+str(round(np.std([p[0]-fit[0]*p[1]-fit[1] for p in pks]),3))+r'$ [keV], $\chi^2_{\nu}='+str(round(sum([(p[0]-fit[0]*p[1]-fit[1])**2/p[2]**2 for p in pks])/float(len(pks)-3),3))+r'$'
		ax[1].text(200,1.45,label)
		ax[1].set_xlabel('Channel Number')
		ax[1].set_ylabel('Fit Residual (keV)')
		f.tight_layout()
		f.subplots_adjust(hspace=0.1)
		if show:
			plt.show()
		else:
			f.savefig('../plots/calibration/energy_calibration.png')
			f.savefig('../plots/calibration/energy_calibration.pdf')
			plt.close()
	def get_efficiency(self,energy,efficiency):
		a,b,c = efficiency[0],efficiency[1],efficiency[2]
		return np.exp(a*np.log(energy)**2+b*np.log(energy)+c)
	def parse_peak_dat(self,l):
		return {'fnm':str(l[0]),'istp':str(l[1]),'dist':float(l[2]),'N':int(l[3]),'unc_N':int(l[4]),'chi2':float(l[5]),'eng':float(l[6]),'I':float(l[7]),'unc_I':float(l[8])}
	def efficiency_calibration(self,show=False):
		guess = [0.0,-0.8,-0.3]
		A0 = {'133BA':39.89E3,'137CS':38.55E3,'152EU':39.29E3,'54MN':37.48E3}
		T0 = {'133BA':dtm.datetime(2009,1,1,0,0,0),'137CS':dtm.datetime(2009,1,1,0,0,0),'152EU':dtm.datetime(2009,1,1,0,0,0),'54MN':dtm.datetime(2009,1,1,0,0,0)}
		itps = {'133BA':isotope('133BA'),'137CS':isotope('137CS'),'152EU':isotope('152EU'),'54MN':isotope('54MN')}
		p = self.pallate
		colr = {1.0:p['k'],5.0:p['b'],10.0:p['r'],15.0:p['y'],18.0:p['p'],24.0:p['g'],29.5:p['gy'],40.0:p['o'],50.0:p['aq'],60.0:p['k']}
		f,ax = plt.subplots()
		E_exp15,N_eff15,sig_N_eff15 = [],[],[]
		for dist in reversed([float(i[0]) for i in self.db.execute('SELECT * FROM calibration')]):
			E_exp,N_eff,sig_N_eff = [],[],[]
			eff_prev = [[float(s) for s in i[2].split(',')] for i in self.db.execute('SELECT * FROM calibration WHERE distance=?',(dist,))][0]
			for pk in [self.parse_peak_dat(i) for i in self.db.execute('SELECT * FROM calibration_peaks') if float(i[10])<float(i[9]) and float(i[5])<300]:
				if pk['I']>=2.0 and pk['dist']==dist and pk['eng']>=50 and pk['unc_N']<0.15*pk['N']:
					lm = itps[pk['istp']].decay_const()
					fl = [i for i in self.spectra if i.filename+'.Spe'==pk['fnm']][0]
					conv = lm/(A0[pk['istp']]*np.exp(-lm*(fl.start_time-T0[pk['istp']]).total_seconds())*(1.0-np.exp(-lm*fl.live_time)))
					E_exp.append(pk['eng'])
					N_eff.append(pk['N']*conv/(0.01*pk['I']))
					sig_N_eff.append(pk['unc_N']*conv/(0.01*pk['I']))
					if abs(N_eff[-1]-self.get_efficiency(E_exp[-1],eff_prev))/sig_N_eff[-1]>7.5:
						del E_exp[-1]
						del N_eff[-1]
						del sig_N_eff[-1]
			if dist==15.0:
				E_exp15,N_eff15,sig_N_eff15 = list(E_exp),list(N_eff),list(sig_N_eff)
			if dist<15.0:
				ratio = 0.5*(N_eff[E_exp.index(834.848)]/N_eff15[E_exp15.index(834.848)]+N_eff[E_exp.index(661.657)]/N_eff15[E_exp15.index(661.657)])
				E_exp,N_eff,sig_N_eff = list(E_exp15),[ratio*i for i in N_eff15],[ratio*i for i in sig_N_eff15]
			efit,unc = curve_fit(lambda x,a,b,c:[np.exp(a*np.log(i)**2+b*np.log(i)+c) for i in x],E_exp,N_eff,p0=guess,sigma=sig_N_eff)
			ax.errorbar(E_exp,N_eff,yerr=sig_N_eff,ls='None',marker='o',color=colr[dist])
			ax.plot(range(50,1450),[np.exp(efit[0]*np.log(i)**2+efit[1]*np.log(i)+efit[2]) for i in range(50,1450)],ls='--',lw=2.0,color=colr[dist])
			ax.text(1475,np.exp(efit[0]*np.log(1475)**2+efit[1]*np.log(1475)+efit[2]),str((int(dist) if int(dist)==dist else dist))+' cm',color=colr[dist])
			self.db.execute('UPDATE calibration SET efficiency_calibration=? WHERE distance=?',(','.join([str(i) for i in efit]),dist))
		self.db_connection.commit()
		ax.set_yscale('log')
		ax.set_xlim((ax.get_xlim()[0],1790))
		ax.set_xlabel('Energy (keV)')
		ax.set_ylabel('Efficiency (a.u.)')
		f.tight_layout()
		if show:
			plt.show()
		else:
			f.savefig('../plots/calibration/efficiency_calibration.pdf')
			f.savefig('../plots/calibration/efficiency_calibration.png')
			plt.close()
	def resolution_calibration(self,show=False):
		f,ax = plt.subplots()
		guess = [0.082]
		pks = [[float(i[0]),float(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT mu,unc_mu,sig,unc_sig FROM calibration_peaks') if float(i[1])<float(i[0])]
		ax.errorbar([np.sqrt(p[0]) for p in pks],[p[2] for p in pks],yerr=[np.sqrt((guess[0]*0.5*p[1]/np.sqrt(p[0]))**2+p[3]**2) for p in pks],ls='None',marker='o',color=self.pallate['r'])
		fit,unc = curve_fit(lambda x,r:[r*i for i in x],[np.sqrt(p[0]) for p in pks],[p[2] for p in pks],p0=guess,sigma=[np.sqrt((guess[0]*0.5*p[1]/np.sqrt(p[0]))**2+p[3]**2) for p in pks])
		for dist in [float(i[0]) for i in self.db.execute('SELECT * FROM calibration')]:
			self.db.execute('UPDATE calibration SET resolution=? WHERE distance=?',(fit[0],dist))
		self.db_connection.commit()
		mu_range = np.arange(min([np.sqrt(p[0]) for p in pks])-4,max([np.sqrt(p[0]) for p in pks])+4,1)
		ax.plot(mu_range,[fit[0]*i for i in mu_range],lw=2.0,color=self.pallate['k'],label=r'$\sigma='+str(round(fit[0],3))+r'\sqrt{bin}$')
		ax.set_xlabel(r'$\sqrt{Bin}$')
		ax.set_ylabel(r'Peak Width $\sigma$ (a.u.)')
		ax.legend(loc=0)
		f.tight_layout()
		if show:
			plt.show()
		else:
			f.savefig('../plots/calibration/resolution.pdf')
			f.savefig('../plots/calibration/resolution.png')
			plt.close()



class ziegler(object):
	def __init__(self):
		self.db_connection = _db_connection
		self.db = _db
		self.dat = {int(i[0]):[float(b) for b in str(i[1]).split(',')] for i in self.db.execute('SELECT * FROM ziegler_dat')}
		self.densities = {str(i[0]):float(i[1]) for i in self.db.execute('SELECT * FROM compounds')}
		self.compounds = {str(i[0]):[(int(h.split(',')[0]),float(h.split(',')[1])) for h in i[2].split(';')] for i in self.db.execute('SELECT * FROM compounds')}
		self.compounds = {cm:[(i[0],i[1]/sum([m[1] for m in self.compounds[cm]])) for i in self.compounds[cm]] for cm in self.compounds}
		self.foils = [{'sample':str(i[1]),'compound':str(i[2]),'density':i[3],'thickness':i[4],'ad':100.0*i[3]*i[4]} for i in self.db.execute('SELECT * FROM foils ORDER BY "order" ASC')]
		air_gap,silicone,kapton = 1.4355,0.043,0.025
		self.stack = [{'sample':self.foils[0]['sample'],'compound':self.foils[0]['compound'],'ad':self.foils[0]['ad'],'density':self.foils[0]['density'],'thickness':self.foils[0]['thickness']}]
		for fl in self.foils[1:]:
			self.stack.append({'sample':'Air','compound':'Air','ad':self.densities['Air']*100.0*air_gap,'density':self.densities['Air'],'thickness':air_gap})
			if any([fl['sample'].startswith(i) for i in ['La','Cu','Al']]):
				self.stack.append({'sample':'Kapton','compound':'Kapton','ad':self.densities['Kapton']*100.0*kapton,'density':self.densities['Kapton'],'thickness':kapton})
				self.stack.append({'sample':'Silicone','compound':'Silicone','ad':self.densities['Silicone']*100.0*silicone,'density':self.densities['Silicone'],'thickness':silicone})
			self.stack.append({'sample':fl['sample'],'compound':fl['compound'],'ad':fl['ad'],'density':fl['density'],'thickness':fl['thickness']})
			if any([fl['sample'].startswith(i) for i in ['La','Cu','Al']]):
				self.stack.append({'sample':'Silicone','compound':'Silicone','ad':self.densities['Silicone']*100.0*silicone,'density':self.densities['Silicone'],'thickness':silicone})
				self.stack.append({'sample':'Kapton','compound':'Kapton','ad':self.densities['Kapton']*100.0*kapton,'density':self.densities['Kapton'],'thickness':kapton})
		self.pallate = plotter().pallate()
	def get_S(self,E,z2):
		# energy E in MeV , stopping power in MeV/(mg/cm2)
		E = 1E3*E/1.00727647
		A,p = self.dat[z2],0.6022140857/self.dat[z2][13]
		if E>=1E3:
			beta_sq = 1.0-1.0/(1.0+E/931478.0)**2
			B0 = np.log(A[6]*beta_sq/(1.0-beta_sq))-beta_sq
			if E<=5E4:
				Y = np.log(E)
				B0 -= A[7]+A[8]*Y+A[9]*Y**2+A[10]*Y**3+A[11]*Y**4
			return (A[5]*p/beta_sq)*B0
		else:
			if E>=10:
				S_low = A[1]*E**0.45
				S_high = A[2]/E*np.log(1.0+A[3]/E + A[4]*E)
				return S_low*S_high*p/(S_low+S_high)
			elif E>0:
				return A[0]*p*E**0.5
			return 0
	def get_dEdx(self,E,cm):
		# E in MeV
		return sum([wt[1]*self.get_S(E,wt[0]) for wt in self.compounds[cm]])
	def solve_dE_stack(self,E0,dp=1.0):
		E = []
		for st in self.stack:
			Ein = E0
			S1 = self.get_dEdx(E0,st['compound'])
			E1 = E0 - dp*st['ad']*S1
			E0 = E0 - 0.5*dp*st['ad']*(S1+self.get_dEdx(E1,st['compound']))
			E.append(0.5*(E0+Ein))
		return E
	def solve_dEdx_AZMC(self,E0=60.0,dE0=0.3,N=1000,dp=1.0):
		foils = [s['sample'] for s in self.foils]
		indices = [n for n,s in enumerate(self.stack) if s['sample'] in foils]
		energies = [[ef for n,ef in enumerate(self.solve_dE_stack(e,dp=dp)) if n in indices] for e in np.random.normal(loc=E0,scale=dE0,size=int(N))]
		histo_dat = [[E[n] for E in energies] for n in range(len(energies[1]))]
		histos = [np.histogram(E,bins='auto') for E in histo_dat]
		dat = [[(foils[n],0.5*(hist[1][m]+hist[1][m+1]),flx) for m,flx in enumerate(hist[0])] for n,hist in enumerate(histos)]
		self.db.execute('DELETE FROM az_tallies')
		for d in dat:
			self.db.executemany('INSERT INTO az_tallies VALUES(?,?,?)',d)
		self.db_connection.commit()
	def get_flux_spectrum(self,foil='La01',table='mcnp'):
		sql = list(self.db.execute('SELECT * FROM '+table+'_tallies WHERE foil=?',(foil,)))
		return [float(i[1]) for i in sql],[float(i[2]) for i in sql]
	def float_to_str(self,f):
		float_string = repr(f)
		if 'e' in float_string:  # detect scientific notation
			digits,exp = float_string.split('e')
			digits,exp = digits.replace('.', '').replace('-', ''),int(exp)
			zero_padding = '0'*(abs(int(exp))-1)  # minus 1 for decimal point in the sci notation
			if exp > 0:
				return '{}{}{}.0'.format('-' if f < 0 else '',digits,zero_padding)
			return '{}0.{}{}'.format('-' if f < 0 else '',zero_padding,digits)
		return float_string
	def generate_mcnp_input(self,E0=60.0,dE0=0.1,N=1e5,dp=1.0):
		ss = '60 Mev Pencil Beam on La Foil Stack mcnp6\n'
		materials = list(set([s['compound'] for s in self.stack]))
		densities = {m:[float(i[1]) for i in self.db.execute('SELECT * FROM compounds WHERE compound=?',(m,))][0] for m in materials}
		for n,foil in enumerate(self.stack):
			dpr = dp #if foil['sample'].startswith('E') else 1.0
			ss += ' '.join([str(n+1),str(materials.index(foil['compound'])+1),'-'+self.float_to_str(round(dpr*densities[foil['compound']],6)),'-500',str(n+1),'-'+str(n+2),'\n'])
		ss += ' '.join(['250',str(materials.index('Air')+1),'-0.000908','-300','500','\n300','0','300','\n','\n1 px 0.0','\n'])
		prev = 0.0
		for n,foil in enumerate(self.stack):
			t = 0.1*foil['thickness']*foil['density']/densities[foil['compound']]
			ss += ' '.join([str(n+2),'px',self.float_to_str(round(t+prev,6)),'\n'])
			prev += t
		ss += ' '.join(['500','rpp','0',str(round(prev,6)),'-1.27','1.27','-1.27','1.27','\n'])
  		ss += ' '.join(['300','so','50','\n','\n'])
  		ss += 'mode n h \n'
  		for n,m in enumerate(materials):
  			ss += 'c '+m+'\n'
  			ln = ''
  			ln += 'm'+str(n+1)+'   '
  			for n,(z,frac) in enumerate(self.compounds[m]):
  				ts = ' '+('0' if z<10 else '')+str(z)+'000. '+self.float_to_str(round(frac,6))
  				ts = ts.replace('01000.','01001.').replace('07000.','07014.').replace('08000.','08016.').replace('13000.','13027.').replace('25000.','25055').replace('57000.','57139.')
  				if len(ln)+len(ts)<80:
  					ln += ts
  				else:
  					ss += ln+' \n'
  					ln = ('     ' if n>0 else '')+ts
  			ss += ln+'\n'
  		ss += '\n'.join(['imp:n   1 {0}r       0','imp:h   1 {0}r       0','sdef  pos=-0.10 0.0 0.0  axs=1 0 0 ext=0 rad=d1 par=h vec=1 0 0 dir=1 erg=d2','']).format(str(len(self.stack)))
  		ss += '\n'.join(['si1 0 0.848101','sp1 -21 1','si2 H {0} 19I {1}','sp2 D 0 0.05 19R','phys:n {2}','phys:h {2} j 0','']).format(str(round(E0-dE0,1)),str(round(E0+dE0,1)),str(round(E0+2*dE0,1)))
  		foils,Np = [s['sample'] for s in self.foils],str(int(N/10**int(np.log(N)/np.log(10))))+'+'+str(int(np.log(N)/np.log(10)))
  		ss += '\n'.join(['prdmp {0} {0} -1','nps {0}','histp {1}','f4:n {1}','e4 1 999I {2}','f14:h {1}','e14 1 999I {2}','']).format(Np,' '.join([str(n+1) for n,s in enumerate(self.stack) if s['sample'] in foils]),str(round(E0+2*dE0,1)))
  		fs = ''
  		for ln in ss.split('\n'):
  			if len(ln)>80:
  				N = max([n for n,s in enumerate(ln) if s==' ' and n<80])
  				fs += ln[:N]+'\n'+'     '+ln[N:]+'\n'
  			else:
  				fs += ln+'\n'
		f = open('mcnp/input.inp','w')
		f.write(fs)
		f.close()
	def run_mcnp(self):
		os.chdir('mcnp')
		rmv = [f for f in list(os.walk('.'))[0][2] if not (f=='input.inp')]
		for f in rmv:
			os.system('rm '+f)
		os.putenv('PATH',os.getenv('PATH')+':'+_mcnp_path+'/MCNP_CODE/bin')
		os.putenv('DATAPATH',_mcnp_path+'/MCNP_DATA')
		os.system('mcnp6 i=input.inp')
		rmv = [f for f in list(os.walk('.'))[0][2] if not (f=='input.inp' or f=='outp')]
		for f in rmv:
			os.system('rm '+f)
		os.chdir('..')
		self.parse_mcnp()
	def parse_mcnp(self):
		E_cutoff = 27
		l14,start,stop = False,False,False
		dat,current_cell = {},0
		for ln in open('mcnp/outp','r').read().split('\n'):
			if stop:
				continue
			if ln.startswith('1tally'):
				if [i for i in ln.split(' ') if i!=''][1]=='14':
					l14 = True
			if ln.startswith(' cell') and l14:
				start = True
				current_cell = int(ln.strip()[4:])
				dat[current_cell] = []
				continue
			if start:
				if ln.startswith(' ==='):
					stop = True
					continue
				l = [i for i in ln.split(' ') if i!='']
				if len(l)==3:
					if l[0]!='total':
						dat[current_cell].append([float(i) for i in l])
		self.db.execute('DELETE FROM mcnp_tallies')
		for foil_id in dat:
			self.db.executemany('INSERT INTO mcnp_tallies VALUES(?,?,?,?)',[(self.stack[foil_id-1]['sample'],d[0],d[1],d[2]) for d in dat[foil_id] if d[0]>E_cutoff])
		self.db_connection.commit()
	def plot_stopping_power_Z(self,Z):
		f,ax = plt.subplots()
		ax.plot(np.arange(0.01,1,0.01).tolist()+np.arange(1,1E4,10).tolist(),[zg.get_S(e,Z) for e in np.arange(0.01,1,0.01).tolist()+np.arange(1,1E4,10).tolist()],lw=2.0,color='#2c3e50',label='dE/dx(Z='+str(Z)+')')
		ax.set_yscale('log')
		ax.set_xscale('log')
		ax.set_ylabel(r'Stopping Power (MeV[mg/cm$^2$]$^{-1}$])')
		ax.set_xlabel('Energy (MeV)')
		ax.legend(loc=0)
		f.tight_layout()
		plt.show()
	def plot_dEdx_AZMC(self,sample='La',table='az',saveplot=True):
		f,ax = plt.subplots()
		once = 2 if table=='both' else 1
		while once>0:
			table = 'mcnp' if once==2 else table
			once -= 1
			for n in range(1,11):
				eng,flux = self.get_flux_spectrum(sample+('0' if n<10 else '')+str(n),table)
				ax.plot(eng,flux,color=self.pallate['k'],ls=('--' if n%2==0 else '-'),label=(table.upper()+' Calculation' if n==10 else None))
				# ax.text(np.average(eng,weights=flux)-(0.25*(n-1)),max(flux)*1.2,sample+('0' if n<10 else '')+str(n),fontsize=10)
				print n,np.average(eng,weights=flux)
			table = 'az' if once==1 else table
		ax.set_yscale('log')
		ax.set_xlabel('Proton Energy (MeV)',fontsize=18)
		ax.set_ylabel('Intensity (a.u.)',fontsize=18)
		ax.legend(loc=0)
		yl,yh = ax.get_ylim()
		if table=='mcnp':
			yl = 2e-6
		ax.set_ylim(yl,4*yh)
		f.tight_layout()
		if saveplot:
			f.savefig('../plots/monitors/'+sample+'_'+table+'_spectrum.png')
			f.savefig('../plots/monitors/'+sample+'_'+table+'_spectrum.pdf')
			plt.close()
		plt.show()


class XS_library(object):
	def __init__(self):
		self.db = _db
		self.db_connection = _db_connection
		self.monitor_ids = ['Cu'+('0' if n<10 else '')+str(n) for n in range(1,11)]
		self.monitor_ids += ['Al'+('0' if n<10 else '')+str(n) for n in range(1,11)]
		self.target_ids = ['La'+('0' if n<10 else '')+str(n) for n in range(1,11)]
		self.daughters = ['134LA','135LA','137CEg','133BAg']
		self.measured_xs = None
		self.pallate = plotter().pallate()
		self.cal_src_err = 0.03
	def save_as_xlsx(self):
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		for n,tt in enumerate(['Foil','Isotope','E [MeV]','unc_E [MeV]','XS [mb]','unc_XS [mb]','TALYS [mb]','EMPIRE [mb]','IAEA rec. [mb]',]):
			ws.cell(row=1,column=n+1,value=tt)
		for n,ln in enumerate([[str(i[0]),str(i[1]),float(i[2]),float(i[3]),float(i[4]),float(i[5])] for i in self.db.execute('SELECT * FROM measured_xs')]):
			for m,i in enumerate(ln):
				ws.cell(row=n+2,column=m+1,value=i)
			if ln[0].startswith('La'):
				ws.cell(row=n+2,column=7,value=str(round(float(self.interpolate(ln[1])(ln[2])),3)))
				ws.cell(row=n+2,column=8,value=str(round(float(self.interpolate(ln[1],table='empire')(ln[2])),3)))
			elif ln[1] in ['58CO','63ZN','62ZN','22NA','24NA']:
				ws.cell(row=n+2,column=9,value=str(round(float(self.interpolate(ln[1],table='monitor')(ln[2])),3)))
		wb.save('../data/cross_sections/la139_XS.xlsx')
	def save_as_csv(self):
		f = open('../data/cross_sections/la139_XS.csv','w')
		ss = ','.join(['#Target','Product','E [MeV]','unc_E [MeV]','XS [mb]','unc_XS [mb]'])+'\n'
		f.write(ss+'\n'.join([','.join([str(i[0]),str(i[1]),str(i[2]),str(i[3]),str(i[4]),str(i[5])]) for i in self.db.execute('SELECT * FROM measured_xs')]))
		f.close()
	def save_as_tex(self):
		f = open('../data/cross_sections/la139_XS.tex','w')
		ss = '\n'.join([r'\appendix',r'\section{Table of Cross Sections}',r'\label{xs_appendix}',r'\ \ ',r'\begin{ruledtabular}',r'\begin{tabular}{ccc}',r'Isotope & E [MeV] & $\sigma$ [mb] \\',''])
		for istp in ['134CE','135CE','137CEm','137CEg','139CE','132CS','133BAm','133BAg','135LA','61CU']:
			ss += '\n'+r'\hline'+'\n'
			ss += r'\multicolumn{3}{l}{'+(r'\textit{Cumulative}' if istp in ['135CE','139CE'] else r'\textit{Independent}')+r'} \\ '+'\n'
			ss += '\n'.join([isotope(i[1]).TeX()+' & '+str(round(i[2],2))+r' $\pm$ '+str(round(i[3],2))+' & '+str(round(i[4],3))+r' $\pm$ '+str(round(i[5],3))+r' \\ ' for i in self.db.execute('SELECT * FROM measured_xs WHERE isotope LIKE ?',(istp,))])
		ss += '\n'.join([r'\end{tabular}',r'\label{table:all_xs}',r'\end{ruledtabular}',r'\ \ ','',''])
		ss += '\n'.join([r'\section{Stack Design}',r'\label{stack_appendix}',r'\ \ ',r'\begin{ruledtabular}',r'\begin{tabular}{cccc}',r'Foil Id & Compound & $\Delta x$ [mm] & $\rho \Delta x$ [mg/cm$^2$] \\',''])
		ss += '\n'.join([str(i[1])+' & '+{'SS_316':'316 SS','La':'La','Al':'Al','Cu':'Cu'}[str(i[2])]+' & '+str(i[4])+' & '+str(round(100.0*float(i[3])*float(i[4]),2))+r' $\pm$ '+str(round(100.0*float(i[5])*float(i[4]),2))+r' \\' for i in self.db.execute('SELECT * FROM foils ORDER BY ?',('order',))])
		ss += '\n'.join([r'\end{tabular}',r'\label{table:stack}',r'\end{ruledtabular}',r'\ \ ',''])
		f.write(ss)
		f.close()
	def exp_smooth(self,ls,alpha=0.3):
		R,RR,b = [ls[0]],[ls[-1]],1.0-alpha
		for i,ii in zip(ls[1:],reversed(ls[:-1])):
			R.append(alpha*i+b*R[-1])
			RR.append(alpha*ii+b*RR[-1])
		return [0.5*(R[n]+r) for n,r in enumerate(reversed(RR))]
	def parse_experiment_peak(self,l):
		cols = ['filename','istp','eng','I','unc_I','N','unc_N','chi2','decay_const','efficiency','count_time','cooling_time']
		return {c:(str(l[n]) if n<2 else float(l[n])) for n,c in enumerate(cols)}
	def get_exfor(self,istp,Emax=100):
		dat = [(float(i[1]),float(i[2]),float(i[3]),float(i[4]),str(i[5])) for i in self.db.execute('SELECT * FROM exfor WHERE isotope=?',(istp,))]
		auth = list(set([i[4] for i in dat]))
		return {a:{'E':[d[0] for d in dat if d[4]==a and d[0]<Emax],'dE':[d[1] for d in dat if d[4]==a and d[0]<Emax],'XS':[d[2] for d in dat if d[4]==a and d[0]<Emax],'dXS':[d[3] for d in dat if d[4]==a and d[0]<Emax]} for a in auth}
	def get_monitor(self,istp):
		return [{'E':float(i[2]),'XS':float(i[3]),'dXS':float(i[4])} for i in self.db.execute('SELECT * FROM monitor_xs WHERE product=?',(istp,))]
	def get_talys(self,istp):
		return [{'E':float(i[1]),'XS':float(i[2])} for i in self.db.execute('SELECT * FROM model_xs WHERE model=? AND isotope=?',('talys',istp))]
	def get_alice(self,istp):
		return [{'E':float(i[1]),'XS':float(i[2])} for i in self.db.execute('SELECT * FROM model_xs WHERE model=? AND isotope=?',('alice',istp))]
	def get_empire(self,istp,smooth=True):
		if smooth:
			xs = [(float(i[1]),float(i[2])) for i in self.db.execute('SELECT * FROM model_xs WHERE model=? AND isotope=? ORDER BY E ASC',('empire',istp))]
			return [{'E':xs[n][0],'XS':i} for n,i in enumerate(self.exp_smooth([i[1] for i in xs]))]
		return [{'E':float(i[1]),'XS':float(i[2])} for i in self.db.execute('SELECT * FROM model_xs WHERE model=? AND isotope=?',('empire',istp))]
	def get_prediction(self,istp):
		return [{'E':float(i[1]),'XS':float(i[2])} for i in self.db.execute('SELECT * FROM xs_prediction WHERE isotope=?',(istp,))]
	def interpolate(self,istp,table='talys'):
		ls = {'talys':self.get_talys,'empire':self.get_empire,'alice':self.get_alice,'monitor':self.get_monitor,'prediction':self.get_prediction}[table](istp)
		return interp1d([i['E'] for i in ls],[i['XS'] for i in ls],bounds_error=False,fill_value='extrapolate')
	def interpolate_unc(self,istp):
		ls = self.get_monitor(istp)
		return interp1d([i['E'] for i in ls],[i['dXS'] for i in ls],bounds_error=False,fill_value=0.0)
	def fit_peaks(self,group='both'):
		for foil_id in {'monitor':self.monitor_ids,'target':self.target_ids,'both':self.monitor_ids+self.target_ids}[group]:
			for spec_fnm in [str(i[0]) for i in self.db.execute('SELECT * FROM all_files WHERE filename LIKE "%'+foil_id+'%"')]:
				spec = spectrum(spec_fnm)
				self.db.execute('DELETE FROM experiment_peaks WHERE filename=?',(spec_fnm,))
				for pk in spec.fit_peaks():
					N = (len(pk['fit'])-2)/3
					for n in range(N):
						w = 'pk_info'
						pkinfo = [spec.filename+'.Spe',pk[w][n]['istp'],pk[w][n]['eng'],pk[w][n]['I'],pk[w][n]['unc_I'],pk[w][n]['N'],pk[w][n]['unc_N'],pk[w][n]['chi2']]
						pkinfo += [pk[w][n]['lm'],spec.get_efficiency(pk[w][n]['eng']),spec.live_time,spec.cooling_time]
						self.db.execute('INSERT INTO experiment_peaks VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',tuple(pkinfo))
				spec.plot_fits(wfit=True,logscale=True,save=True,zoom=True)
				self.db_connection.commit()
	def filter_outliers(self,A,sigA,T_cool,A0,lm):
		good_pts = []
		for n,t in enumerate(T_cool):
			if A0 is None:
				if A[n]*np.exp(lm*t)<1e6 and A[n]>sigA[n]:
					good_pts.append(n)
			else:
				if abs(A0*np.exp(-lm*t)-A[n])/sigA[n]<2.5 and A[n]>sigA[n] and abs(A0*np.exp(-lm*t)-A[n])/A0*np.exp(-lm*t)<3:
					good_pts.append(n)
		return [A[n] for n in good_pts],[sigA[n] for n in good_pts],[T_cool[n] for n in good_pts]
	def fit_initial_activities(self,group='both',saveplot=True,logscale=False):
		for foil_id in (self.monitor_ids+self.target_ids if group=='both' else {'monitor':self.monitor_ids,'target':self.target_ids}[group]):
			for itp in [str(i[1]).split(',') for i in self.db.execute('SELECT * FROM sample_map WHERE sample=?',(foil_id[:2],))][0]:
				self.db.execute('DELETE FROM initial_activities WHERE foil=? AND isotope=?',(foil_id,itp))
				if itp in self.daughters or (itp=='22NA' and foil_id[:2]=='Cu') or (itp=='22NA' and foil_id=='Al05'):
					continue
				ip = isotope(itp)
				peaks = [self.parse_experiment_peak(l) for l in self.db.execute('SELECT * FROM experiment_peaks WHERE filename LIKE "%'+foil_id+'%" AND isotope=?',(itp,))]
				A = [(p['decay_const']*p['N'])/((1.0-np.exp(-p['decay_const']*p['count_time']))*0.01*p['I']*p['efficiency']) for p in peaks]
				sigA = [np.sqrt(p['unc_N']**2*(A[n]/p['N'])**2+p['unc_I']**2*(A[n]/p['I'])**2+ip.unc_decay_const()**2*(A[n]/p['decay_const'])) for n,p in enumerate(peaks)]
				T_cool = [p['cooling_time'] for p in peaks]
				lm = ip.decay_const()
				A,sigA,T_cool = self.filter_outliers(A,sigA,T_cool,None,lm)
				if len(A)==0:
					continue
				for itr in range(2):
					decay = lambda x,A0:[A0*np.exp(-lm*t) for t in x]
					fit,unc = curve_fit(decay,T_cool,A,p0=[A[0]*np.exp(lm*T_cool[0])],sigma=sigA)
					A,sigA,T_cool = self.filter_outliers(A,sigA,T_cool,fit[0],lm)
				if len(A)>0:
					fit,unc = curve_fit(decay,T_cool,A,p0=[A[0]*np.exp(lm*T_cool[0])],sigma=sigA)
				wts = [np.exp(lm*t)/np.sqrt(sigA[n]**2*(A[n]-fit[0]*np.exp(lm*t)-1e-9)**2) for n,t in enumerate(T_cool)]
				sm = sum(wts)
				wts = [i/sm for i in wts]
				# unc = [[(unc[0][0] if len(A)>1 else (sigA[0]*np.exp(lm*T_cool[0]))**2)]]
				unc = [[(sum([sigA[n]**2*w**2 for n,w in enumerate(wts)]))]]
				self.db.execute('INSERT INTO initial_activities VALUES(?,?,?,?)',(foil_id,itp,fit[0],np.sqrt(unc[0][0]+ip.unc_decay_const()**2*(fit[0]/lm)**2+(self.cal_src_err*fit[0])**2)))
				if saveplot:
					f,ax = plt.subplots()
					Trange = np.arange(0,max(T_cool)/3600.0+0.1,0.1)
					ax.errorbar([t/3600.0 for t in T_cool],[a/1e3 for a in A],yerr=[a/1e3 for a in sigA],ls='None',marker='o',color=self.pallate['k'],label='Photopeak Activity')
					ax.plot(Trange,[0.001*fit[0]*np.exp(-lm*3600.0*t) for t in Trange],color=self.pallate['gy'],label='Exponential Fit')
					ax.plot(Trange,[0.001*(fit[0]+np.sqrt(unc[0][0]))*np.exp(-lm*3600.0*t) for t in Trange],ls='--',color=self.pallate['gy'],label=r'$\pm 1\sigma_{fit}$')
					ax.plot(Trange,[0.001*(fit[0]-np.sqrt(unc[0][0]))*np.exp(-lm*3600.0*t) for t in Trange],ls='--',color=self.pallate['gy'])
					if logscale:
						ax.set_yscale('log')
					ax.set_xlabel('Cooling Time (h)')
					ax.set_ylabel('Activity (kBq)')
					ax.set_title(ip.TeX()+' Activity in '+foil_id)
					ax.legend(loc=0)
					f.tight_layout()
					f.savefig('../plots/decay_curves/'+foil_id+'_'+itp+'.png')
					f.savefig('../plots/decay_curves/'+foil_id+'_'+itp+'.pdf')
					plt.close()
		self.db_connection.commit()
	def filter_daughter_outliers(self,A,sigA,T_c,T_m,peaks,decay_func=None,p0=None):
		good_pts = []
		if decay_func is None:
			for n,t in enumerate(T_c):
				if A[n]>sigA[n]:
					good_pts.append(n)
		else:
			for n,t in enumerate(T_c):
				if abs(A[n]-decay_func([t],p0)[0])/sigA[n]<2.5 and abs(A[n]-decay_func([t],p0)[0])/decay_func([t],p0)[0]<3:
					good_pts.append(n)
		return [A[n] for n in good_pts],[sigA[n] for n in good_pts],[T_c[n] for n in good_pts],[T_m[n] for n in good_pts],[peaks[n] for n in good_pts]
	def fit_parent_daughter_activities(self,saveplot=True,logscale=False):
		for foil_id in self.target_ids:
			parents = {'134LA':'134CE','135LA':'135CE','137CEg':'137CEm','133BAg':'133BAm'}
			for itp in self.daughters:
				if itp=='134LA':
					continue
				self.db.execute('DELETE FROM initial_activities WHERE foil=? AND isotope=?',(foil_id,itp))
				ip = isotope(itp)
				parent_ip = isotope(parents[itp])
				lm_d,lm_p = ip.decay_const(),parent_ip.decay_const()
				peaks = [self.parse_experiment_peak(l) for l in self.db.execute('SELECT * FROM experiment_peaks WHERE filename LIKE "%'+foil_id+'%" AND isotope=?',(itp,))]
				if len(peaks)==0:
					continue
				try:
					A0_p,unc_A0_p = [(float(i[2]),float(i[3])) for i in self.db.execute('SELECT * FROM initial_activities WHERE foil=? AND isotope=?',(foil_id,parent_ip.name()))][0]
				except:
					continue
				E_proton = [float(i[1]) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil=?',(foil_id,))][0]
				xsp,xsd = self.interpolate(parent_ip.name(),table='prediction'),self.interpolate(ip.name(),table='prediction')
				xsp,xsd = max((xsp(E_proton),1e-5)),max((xsd(E_proton),1e-5))
				fit,unc = [xsd*A0_p/xsp],[[(xsd*unc_A0_p/xsp)**2]]
				T_c,T_m = [p['cooling_time'] for p in peaks],[p['count_time'] for p in peaks]
				for itr in range(3):
					e = np.exp
					A = [(p['N']/(0.01*p['I']*p['efficiency']))*((A0_p*lm_d/(lm_d-lm_p))*(e(-lm_p*T_c[n])-e(-lm_d*T_c[n]))+fit[0]*e(-lm_d*T_c[n]))/((A0_p/(lm_d-lm_p))*((e(-lm_d*(T_m[n]+T_c[n]))-e(-lm_d*T_c[n]))-(lm_d/lm_p)*(e(-lm_p*(T_m[n]+T_c[n]))-e(-lm_p*T_c[n])))-(fit[0]/lm_d)*(e(-lm_d*(T_c[n]+T_m[n]))-e(-lm_d*T_c[n]))) for n,p in enumerate(peaks)]
					sigA = [np.sqrt(p['unc_N']**2*(A[n]/p['N'])**2+p['unc_I']**2*(A[n]/p['I'])**2) for n,p in enumerate(peaks)]
					decay_func = lambda x,A0_d:[A0_p*(lm_d/(lm_d-lm_p))*(e(-lm_p*tc)-e(-lm_d*tc))+A0_d*e(-lm_d*tc) for tc in x]
					A,sigA,T_c,T_m,peaks = self.filter_daughter_outliers(A,sigA,T_c,T_m,peaks)
					if len(A)==0:
						continue
					fit,unc = curve_fit(decay_func,T_c,A,p0=fit,sigma=sigA,bounds=([0],[np.inf]))
					A,sigA,T_c,T_m,peaks = self.filter_daughter_outliers(A,sigA,T_c,T_m,peaks,decay_func,fit)
				if len(A)==0:
					continue
				# unc = [[(unc[0][0] if len(A)>1 else sigA[0]**2*(decay_func([T_c[0]],fit[0])[0]/A[0])**2)]]
				wts = [np.exp(lm_d*T_c[n])/np.sqrt(sigA[n]**2*(A[n]-a-1e-9)**2) for n,a in enumerate(decay_func(T_c,fit[0]))]
				sm = sum(wts)
				wts = [i/sm for i in wts]
				unc = [[(sum([sigA[n]**2*w**2 for n,w in enumerate(wts)])+unc_A0_p**2*(fit[0]/A0_p)**2)]]
				self.db.execute('INSERT INTO initial_activities VALUES(?,?,?,?)',(foil_id,itp,fit[0],np.sqrt(unc[0][0]+ip.unc_decay_const()**2*(fit[0]/lm_d)**2+parent_ip.unc_decay_const()**2*(fit[0]/lm_p)**2+(self.cal_src_err*fit[0])**2)))
				if saveplot:
					f,ax = plt.subplots()
					Trange = np.arange(0,max(T_c)/3600.0+0.1,0.1)
					ax.errorbar([t/3600.0 for t in T_c],[0.001*a for a in A],yerr=[0.001*a for a in sigA],ls='None',marker='o',color=self.pallate['k'],label='Photopeak Activity')
					ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0])],color=self.pallate['gy'],label='Exponential Fit')
					ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0]+np.sqrt(unc[0][0]))],ls='--',color=self.pallate['gy'],label=r'$\pm 1\sigma_{fit}$')
					ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0]-np.sqrt(unc[0][0]))],ls='--',color=self.pallate['gy'])
					if logscale:
						ax.set_yscale('log')
					ax.set_xlabel('Cooling Time (h)')
					ax.set_ylabel('Activity (kBq)')
					ax.set_title(ip.TeX()+' Activity in '+foil_id)
					ax.legend(loc=0)
					f.tight_layout()
					f.savefig('../plots/decay_curves/'+foil_id+'_'+itp+'.png')
					f.savefig('../plots/decay_curves/'+foil_id+'_'+itp+'.pdf')
					plt.close()
		self.db_connection.commit()
	def fit_134CE_activity(self,saveplot=True,logscale=False):
		for foil_id in self.target_ids:
			self.db.execute('DELETE FROM initial_activities WHERE foil=? AND isotope=?',(foil_id,'134LA'))
			ip = isotope('134LA')
			parent_ip = isotope('134CE')
			lm_d,lm_p = ip.decay_const(),parent_ip.decay_const()
			peaks = [self.parse_experiment_peak(l) for l in self.db.execute('SELECT * FROM experiment_peaks WHERE filename LIKE "%'+foil_id+'%" AND isotope=?',('134LA',))]
			if len(peaks)==0:
				continue
			T_c,T_m = [p['cooling_time'] for p in peaks],[p['count_time'] for p in peaks]
			for itr in range(2):
				e = np.exp
				A = [(p['N']*lm_d/(0.01*p['I']*p['efficiency']))*(e(-lm_p*T_c[n])-e(-lm_d*T_c[n]))/((e(-lm_d*(T_m[n]+T_c[n]))-e(-lm_d*T_c[n]))-(lm_d/lm_p)*(e(-lm_p*(T_m[n]+T_c[n]))-e(-lm_p*T_c[n]))) for n,p in enumerate(peaks)]
				sigA = [np.sqrt(p['unc_N']**2*(A[n]/p['N'])**2+p['unc_I']**2*(A[n]/p['I'])**2) for n,p in enumerate(peaks)]
				decay_func = lambda x,A0_p:[A0_p*(lm_d/(lm_d-lm_p))*(e(-lm_p*tc)-e(-lm_d*tc)) for tc in x]
				A,sigA,T_c,T_m,peaks = self.filter_daughter_outliers(A,sigA,T_c,T_m,peaks)
				if len(A)==0:
					continue
				fit,unc = curve_fit(decay_func,T_c,A,p0=[1e3],sigma=sigA,bounds=([0],[np.inf]))
				A,sigA,T_c,T_m,peaks = self.filter_daughter_outliers(A,sigA,T_c,T_m,peaks,decay_func,fit)
			if len(A)==0:
				continue
			# unc = [[(unc[0][0] if len(A)>1 else sigA[0]**2*(decay_func([T_c[0]],fit[0])[0]/A[0])**2)]]
			wts = [1.0/np.sqrt(sigA[n]**2*(A[n]-a-1e-9)**2) for n,a in enumerate(decay_func(T_c,fit[0]))]
			sm = sum(wts)
			wts = [i/sm for i in wts]
			unc = [[(sum([sigA[n]**2*w**2 for n,w in enumerate(wts)]))]]
			self.db.execute('INSERT INTO initial_activities VALUES(?,?,?,?)',(foil_id,'134LA',fit[0],np.sqrt(unc[0][0]+ip.unc_decay_const()**2*(fit[0]/lm_d)**2+parent_ip.unc_decay_const()**2*(fit[0]/lm_p)**2+(self.cal_src_err*fit[0])**2)))
			if saveplot:
				f,ax = plt.subplots()
				Trange = np.arange(0,max(T_c)/3600.0+0.1,0.1)
				ax.errorbar([t/3600.0 for t in T_c],[0.001*a for a in A],yerr=[0.001*a for a in sigA],ls='None',marker='o',color=self.pallate['k'],label='Photopeak Activity')
				ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0])],color=self.pallate['gy'],label='Exponential Fit')
				ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0]+np.sqrt(unc[0][0]))],ls='--',color=self.pallate['gy'],label=r'$\pm 1\sigma_{fit}$')
				ax.plot(Trange,[0.001*a for a in decay_func([3600.0*t for t in Trange],fit[0]-np.sqrt(unc[0][0]))],ls='--',color=self.pallate['gy'])
				if logscale:
					ax.set_yscale('log')
				ax.set_xlabel('Cooling Time (h)')
				ax.set_ylabel('Activity (kBq)')
				ax.set_title(ip.TeX()+' Activity in '+foil_id)
				ax.legend(loc=0)
				f.tight_layout()
				f.savefig('../plots/decay_curves/'+foil_id+'_134LA.png')
				f.savefig('../plots/decay_curves/'+foil_id+'_134LA.pdf')
				plt.close()
		self.db_connection.commit()
	def calculate_cross_sections(self,La34=True):
		self.db.execute('DELETE FROM measured_xs')
		self.measured_xs = []
		irr_time = 5844.0
		for istp in ['132CS','133BAm','133BAg','134CE','135CE','135LA','137CEg','137CEm','139CE','58CO','61CU','62ZN','63ZN','22NA','24NA']:
			A0 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=?',(istp,)) if (str(i[0])[:2]=='Al' if istp in ['22NA','24NA'] else True)]
			if istp=='134CE':
				if La34:
					A0 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=?',('134LA',))]
			ip = isotope(istp)
			for idx,fl in enumerate([i[0] for i in A0]):
				if fl=='La03' or int(fl[2:])-1 in {'134CE':[5,6,7,8,9],'133BAg':[0,4,7,8,9],'132CS':[5,6,7,8,9],'22NA':[1],'24NA':[0],'139CE':[8],'135LA':[4,8,9]}.setdefault(istp,[]):
					continue
				if istp in ['22NA','24NA'] and fl[:2]=='Al':
					Al_corr = [(float(i[2]),float(i[3])) for i in self.db.execute('SELECT * FROM initial_activities WHERE foil LIKE "%'+fl[2:]+'%" AND isotope=?',(istp,)) if not str(i[0])[:2]=='Al']
					mean = np.average([i[0] for i in Al_corr],weights=[1.0/i[1] for i in Al_corr])
					A0[idx][2] -= mean
					A0[idx][3] = np.sqrt(A0[idx][3]**2+np.average([i[1]**2 for i in Al_corr],weights=[1.0/i[1] for i in Al_corr]))
				density,thickness,unc_density = [(float(i[3]),float(i[4]),float(i[5])) for i in self.db.execute('SELECT * FROM foils WHERE sample=?',(fl,))][0]
				Z = [int(str(i[2]).split(',')[0]) for i in self.db.execute('SELECT * FROM compounds WHERE compound=?',(fl[:2],))][0]
				gm_per_mol = [float(str(i[1]).split(',')[13]) for i in self.db.execute('SELECT * FROM ziegler_dat WHERE z=?',(Z,))][0]
				E,dE,Ip,unc_Ip = [(float(i[1]),float(i[2]),float(i[3]),float(i[4])) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil=?',(fl,))][0]
				self.measured_xs.append({'foil':fl,'istp':istp,'E':E,'dE':dE})
				sigma = 1e3*A0[idx][2]/((Ip/1.602e-10)*(0.1*density*thickness*6.022e-1/gm_per_mol)*(1.0-np.exp(-ip.decay_const()*irr_time)))
				self.measured_xs[-1]['sigma'] = sigma
				self.measured_xs[-1]['unc_sigma'] = np.sqrt(A0[idx][3]**2*(sigma/A0[idx][2])**2+unc_density**2*(sigma/density)**2+unc_Ip**2*(sigma/Ip)**2)
		self.db.executemany('INSERT INTO measured_xs VALUES(?,?,?,?,?,?)',[(xs['foil'],xs['istp'],xs['E'],xs['dE'],xs['sigma'],xs['unc_sigma']) for xs in self.measured_xs])
		self.db_connection.commit()
	def plot_cross_sections(self,saveplot=True,prediction=True,La34=True,plot_xs=None):
		if self.measured_xs is None:
			self.calculate_cross_sections(La34)
		xs_list = list(set([i['istp'] for i in self.measured_xs])) if plot_xs is None else [plot_xs]
		foil_map = {i[1]:i[0] for i in list(set([(str(l[0][:2]),str(l[1])) for l in self.db.execute('SELECT * FROM measured_xs')]))}
		for istp in xs_list:
			ip = isotope(istp)
			xs = [i for i in self.measured_xs if i['istp']==istp and (i['foil'][:2]=='Al' if istp in ['22NA','24NA'] else True)]
			f,ax = plt.subplots()
			ms = 6.0 if istp in ['58CO','61CU'] else 4.5
			if prediction:
				if istp in ['61CU']:
					x4 = self.get_exfor(istp)
					for n,a in enumerate(x4):
						ax.errorbar(x4[a]['E'],x4[a]['XS'],xerr=x4[a]['dE'],yerr=x4[a]['dXS'],ls='None',marker='o',color=self.pallate['gy'],label=a)
				elif istp in ['58CO','63ZN','62ZN','22NA','24NA']:
					Erange = np.arange(min([i['E'] for i in xs])-1.0,max([i['E'] for i in xs]),0.1)
					ax.plot(Erange,self.interpolate(istp,'monitor')(Erange),color=self.pallate['b'],label='IAEA Rec. Value')
				else:
					Erange = np.arange(min([i['E'] for i in xs])-1.0,max([i['E'] for i in xs]),0.1)
					scl_talys = 0.1 if istp=='134CE' else 1.0
					scl_empire = 0.05 if istp=='132CS' else 1.0
					ax.plot(Erange,scl_talys*self.interpolate(istp,'talys')(Erange),color=self.pallate['b'],label=('0.1x ' if istp=='134CE' else '')+'TALYS Calculation',lw=2.0)
					ax.plot(Erange,scl_empire*self.interpolate(istp,'empire')(Erange),color=self.pallate['r'],ls='--',label=('0.05x ' if istp=='132CS' else '')+'EMPIRE Calculation',lw=2.0)
					# ax.plot(Erange,scl_talys*self.interpolate(istp,'alice')(Erange),color=self.pallate['p'],ls='-.',label=('0.1x ' if istp=='134CE' else '')+'ALICE Calculation',lw=2.0)
			ax.errorbar([i['E'] for i in xs],[i['sigma'] for i in xs],xerr=[i['dE'] for i in xs],yerr=[i['unc_sigma'] for i in xs],ms=ms,ls='None',marker='o',color=self.pallate['k'],label=r'Measured $\sigma$'+'\n(this work)') #,elinewidth=3.0
			ax.set_xlabel('Energy (MeV)')#,fontsize=18
			ax.set_ylabel('Cross Section (mb)')#,fontsize=18
			ax.set_title(r'$^{nat}$'+foil_map[istp]+'(p,x)'+ip.TeX()+' Cross Section')
			ax.legend(loc=0,fontsize=(10 if istp in ['58CO','61CU'] and prediction else None),borderaxespad=0.75)
			f.tight_layout()
			if saveplot:
				f.savefig('../plots/cross_sections/'+istp+('' if La34 else '_CE_fit')+('' if prediction else '_only')+'.png')
				f.savefig('../plots/cross_sections/'+istp+('' if La34 else '_CE_fit')+('' if prediction else '_only')+'.pdf')
				plt.close()
			else:
				plt.show()
	def ratios(self):
		Co58 = self.interpolate('58CO','monitor')
		Zn62 = self.interpolate('62ZN','monitor')
		Zn63 = self.interpolate('63ZN','monitor')
		unc_Co58 = self.interpolate_unc('58CO')
		unc_Zn62 = self.interpolate_unc('62ZN')
		unc_Zn63 = self.interpolate_unc('63ZN')
		f,ax = plt.subplots(1,3,figsize=(19.2,4.8))
		E_range = np.arange(30,60,0.05)
		itrp = False
		E_meas = [float(i[1]) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil LIKE "%Cu%" ORDER BY foil ASC')]
		unc_E_meas = [float(i[2]) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil LIKE "%Cu%" ORDER BY foil ASC')]
		

		A0 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('58CO',))]
		A1 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('62ZN',))]
		A_ratio = [a[2]/A1[n][2] for n,a in enumerate(A0)]
		unc_A_ratio = [np.sqrt((a[3]/A1[n][2])**2+(A1[n][3]*a[2]/A1[n][2]**2)**2) for n,a in enumerate(A0)]
		if itrp:
			E_meas = [50.86,48.61,46.99,44.74,39.13,37.53,34.73]
			unc_E_meas = [2.1,2.1,2.5,2.1,2.1,1.0,0.4]
		R = (1.0-np.exp(-5844.0*isotope('58CO').decay_const()))/(1.0-np.exp(-5844.0*isotope('62ZN').decay_const()))
		ax[0].plot(E_range,R*Co58(E_range)/Zn62(E_range),label='Predicted (IAEA)',color=self.pallate['gy'],lw=1.8)
		upper = [(R*Co58(e)/Zn62(e))+R*np.sqrt((unc_Co58(e)/Zn62(e))**2+(unc_Zn62(e)*Co58(e)/Zn62(e)**2)**2) for e in E_range]
		lower = [(R*Co58(e)/Zn62(e))-R*np.sqrt((unc_Co58(e)/Zn62(e))**2+(unc_Zn62(e)*Co58(e)/Zn62(e)**2)**2) for e in E_range]
		ax[0].fill_between(E_range,upper,lower,alpha=0.4,color=self.pallate['gy'])
		if itrp:
			ax[0].errorbar(E_meas,A_ratio[3:],xerr=unc_E_meas,yerr=unc_A_ratio[3:],color=self.pallate['k'],ls='None',marker='o',label='Measured')
		else:
			ax[0].errorbar(E_meas,A_ratio,xerr=unc_E_meas,yerr=unc_A_ratio,color=self.pallate['k'],ls='None',marker='o',label='Measured')
		ax[0].set_ylabel(r'$^{58}$Co/$^{62}$Zn Activity Ratio')
		ax[0].set_xlabel('Energy (MeV)')
		ax[0].legend(loc=0)

		A0 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('62ZN',))]
		A1 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('63ZN',))]
		A_ratio = [a[2]/A1[n][2] for n,a in enumerate(A0)]
		unc_A_ratio = [np.sqrt((a[3]/A1[n][2])**2+(A1[n][3]*a[2]/A1[n][2]**2)**2) for n,a in enumerate(A0)]
		if itrp:
			E_meas = [53.43,51.31,50.73,49.68,48.26,45.17,42.73,40.04,35.76]
			unc_E_meas = [6.3,2.4,1.5,1.7,0.8,0.9,0.8,1.0,1.3]
		R = (1.0-np.exp(-5844.0*isotope('62ZN').decay_const()))/(1.0-np.exp(-5844.0*isotope('63ZN').decay_const()))
		ax[1].plot(E_range,R*Zn62(E_range)/Zn63(E_range),label='Predicted (IAEA)',color=self.pallate['gy'],lw=1.8)
		upper = [(R*Zn62(e)/Zn63(e))+R*np.sqrt((unc_Zn62(e)/Zn63(e))**2+(unc_Zn63(e)*Zn62(e)/Zn63(e)**2)**2) for e in E_range]
		lower = [(R*Zn62(e)/Zn63(e))-R*np.sqrt((unc_Zn62(e)/Zn63(e))**2+(unc_Zn63(e)*Zn62(e)/Zn63(e)**2)**2) for e in E_range]
		ax[1].fill_between(E_range,upper,lower,alpha=0.4,color=self.pallate['gy'])
		if itrp:
			ax[1].errorbar(E_meas,[a for n,a in enumerate(A_ratio) if n!=8],xerr=unc_E_meas,yerr=[a for n,a in enumerate(unc_A_ratio) if n!=8],color=self.pallate['k'],ls='None',marker='o',label='Measured')
		else:
			ax[1].errorbar(E_meas,A_ratio,xerr=unc_E_meas,yerr=unc_A_ratio,color=self.pallate['k'],ls='None',marker='o',label='Measured')
		ax[1].set_ylabel(r'$^{62}$Zn/$^{63}$Zn Activity Ratio')
		ax[1].set_xlabel('Energy (MeV)')
		ax[1].legend(loc=0)

		A0 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('58CO',))]
		A1 = [[str(i[0]),str(i[1]),float(i[2]),float(i[3])] for i in self.db.execute('SELECT * FROM initial_activities WHERE isotope=? ORDER BY foil ASC',('63ZN',))]
		A_ratio = [a[2]/A1[n][2] for n,a in enumerate(A0)]
		unc_A_ratio = [np.sqrt((a[3]/A1[n][2])**2+(A1[n][3]*a[2]/A1[n][2]**2)**2) for n,a in enumerate(A0)]
		if itrp:
			E_meas = [45.56,42.40,41.52,39.57,37.11,34.33]
			unc_E_meas = [10.0,3.0,2.3,0.6,0.5,0.4]
		R = (1.0-np.exp(-5844.0*isotope('58CO').decay_const()))/(1.0-np.exp(-5844.0*isotope('63ZN').decay_const()))
		ax[2].plot(E_range,R*Co58(E_range)/Zn63(E_range),label='Predicted (IAEA)',color=self.pallate['gy'],lw=1.8)
		upper = [(R*Co58(e)/Zn63(e))+R*np.sqrt((unc_Co58(e)/Zn63(e))**2+(unc_Zn63(e)*Co58(e)/Zn63(e)**2)**2) for e in E_range]
		lower = [(R*Co58(e)/Zn63(e))-R*np.sqrt((unc_Co58(e)/Zn63(e))**2+(unc_Zn63(e)*Co58(e)/Zn63(e)**2)**2) for e in E_range]
		ax[2].fill_between(E_range,upper,lower,alpha=0.4,color=self.pallate['gy'])
		if itrp:
			ax[2].errorbar(E_meas,A_ratio[4:],xerr=unc_E_meas,yerr=unc_A_ratio[4:],color=self.pallate['k'],ls='None',marker='o',label='Measured')
		else:
			ax[2].errorbar(E_meas,A_ratio,xerr=unc_E_meas,yerr=unc_A_ratio,color=self.pallate['k'],ls='None',marker='o',label='Measured')
		ax[2].set_ylabel(r'$^{58}$Co/$^{63}$Zn Activity Ratio')
		ax[2].set_xlabel('Energy (MeV)')
		ax[2].legend(loc=0)

		
		f.tight_layout()
		plt.show()
	def plot_energy_index(self):
		Eng60 = True
		f,ax = plt.subplots()
		E_meas = [50.86,48.61,46.99,44.74,39.13,37.53,34.73]
		unc_E_meas = [2.1,2.1,2.5,2.1,2.1,1.0,0.4]
		idx = [4,5,6,7,8,9,10]
		ax.errorbar(idx,E_meas,yerr=unc_E_meas,ls='None',marker='o',color=self.pallate['b'],capsize=5.0,label='Activity Ratio Energies')
		E_meas = [53.43,51.31,50.73,49.68,48.26,45.17,42.73,40.04,35.76]
		unc_E_meas = [6.3,2.4,1.5,1.7,0.8,0.9,0.8,1.0,1.3]
		idx = [1,2,3,4,5,6,7,8,10]
		ax.errorbar(idx,E_meas,yerr=unc_E_meas,ls='None',marker='o',color=self.pallate['k'],capsize=5.0)
		E_meas = [45.56,42.40,41.52,39.57,37.11,34.33]
		unc_E_meas = [10.0,3.0,2.3,0.6,0.5,0.4]
		idx = [5,6,7,8,9,10]
		ax.errorbar(idx,E_meas,yerr=unc_E_meas,ls='None',marker='o',color=self.pallate['r'],capsize=5.0)
		# print [round(i[1],2) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil LIKE "%Cu%" ORDER BY foil ASC')]
		
		if Eng60:
			E_mcnp_opt = [58.24, 56.32, 54.33, 52.3, 50.22, 48.06, 44.87, 41.49, 37.9, 33.99]
			E_az_opt = [58.13, 56.18, 54.19, 52.14, 50.02, 47.85, 44.75, 41.46, 37.99, 34.2]
			E_mcnp_def = [58.7, 57.28, 55.87, 54.38, 52.87, 51.37, 49.12, 46.82, 44.43, 41.92]
			E_az_def = [58.42, 56.78, 55.1, 53.38, 51.61, 49.81, 47.27, 44.61, 41.84, 38.89]
		else:
			E_mcnp_def = [55.65, 54.18, 52.68, 51.12, 49.56, 47.95, 45.6, 43.14, 40.61, 37.89]
			E_az_opt = [55.35, 53.63, 51.88, 50.07, 48.22, 46.32, 43.62, 40.78, 37.79, 34.58]
			E_mcnp_opt = [55.44, 53.76, 52.0, 50.21, 48.39, 46.48, 43.72, 40.79, 37.7, 34.36]
		ax.plot(range(1,11),E_mcnp_opt,lw=1.8,color=self.pallate['r'],ls='-.',label='MCNP Optimum')
		ax.plot(range(1,11),E_az_opt,lw=1.8,color=self.pallate['k'],ls='--',label='AZ Optimum')
		ax.plot(range(1,11),E_mcnp_def,lw=1.8,color=self.pallate['p'],ls=':',label='MCNP Default')
		if Eng60:
			ax.plot(range(1,11),E_az_def,lw=1.8,color=self.pallate['gy'],ls='-',label='AZ Default')
		
		ax.set_ylabel('Energy (MeV)')
		ax.set_xlabel('Foil Index')
		ax.legend(loc=0)
		f.tight_layout()
		plt.show()

class monitors(object):
	def __init__(self):
		self.db = _db
		self.db_connection = _db_connection
		self.zg = ziegler()
		self.xslib = XS_library()
		self.foil_ids = ['Cu'+('0' if n<10 else '')+str(n) for n in range(1,11)]
		self.foil_ids += ['Al'+('0' if n<10 else '')+str(n) for n in range(1,11)]
		self.sample_map = {'Cu':['62ZN','63ZN','58CO'],'Al':['22NA','24NA']}
		self.monitor_channels = None
		self.pallate = plotter().pallate()
	def simple_regression(self,x,y):
		xb,yb = np.average(x),np.average(y)
		m = sum([(i-xb)*(y[n]-yb) for n,i in enumerate(x)])/sum([(i-xb)**2 for i in x])
		return m,yb-m*xb
	def get_eng_dist(self,foil='Cu01',table='mcnp'):
		eng,hist = self.zg.get_flux_spectrum(foil,table)
		if sum(hist)==0:
			return None,None
		mean = np.average(eng,weights=hist)
		return mean,np.sqrt(np.average((eng-mean)**2,weights=hist))
	def get_A0(self,foil_id,istp):
		A0 = [(float(i[2]),float(i[3])) for i in self.db.execute('SELECT * FROM initial_activities WHERE foil=? AND isotope=?',(foil_id,istp))]
		if len(A0)==0:
			return None,None
		A0,sigA0 = A0[0][0],A0[0][1]
		if foil_id.startswith('Al'):
			Al_corr = [(float(i[2]),float(i[3])) for i in self.db.execute('SELECT * FROM initial_activities WHERE foil LIKE "%'+foil_id[2:]+'%" AND isotope=?',(istp,)) if not str(i[0])[:2]=='Al']
			mean = np.average([i[0] for i in Al_corr],weights=[1.0/i[1] for i in Al_corr])
			return A0-mean,np.sqrt(sigA0**2+np.average([i[1]**2 for i in Al_corr],weights=[1.0/i[1] for i in Al_corr]))
		return A0,sigA0
	def plot_Al_correction(self,istp='22NA',saveplot=True):
		f,ax = plt.subplots()
		clr = {'Al':self.pallate['r'],'La':self.pallate['gy'],'Cu':self.pallate['k']}
		A0 = {f:[(str(i[0]),str(i[1]),1e-3*float(i[2]),1e-3*float(i[3])) for i in self.db.execute('SELECT * FROM initial_activities WHERE foil LIKE "%'+f+'%" AND isotope=? ORDER BY foil ASC',(istp,))] for f in ['Al','Cu','La']}
		E0 = {f:[(str(i[0]),float(i[1]),float(i[2])) for i in self.db.execute('SELECT * FROM monitor_data WHERE foil LIKE "%'+f+'%" ORDER BY foil ASC')] for f in ['Al','Cu','La']}
		for fl in A0:
			if istp=='22NA' and fl.startswith('Cu'):
				continue
			A0_foils = [i[0] for i in A0[fl]]
			E0_filtered = [(i[1],i[2]) for i in E0[fl] if i[0] in A0_foils]
			ax.errorbar([i[0] for i in E0_filtered],[i[2] for i in A0[fl]],xerr=[i[1] for i in E0_filtered],yerr=[i[3] for i in A0[fl]],ls='None',marker='o',color=clr[fl],label=fl+' Foils')
		ax.set_xlabel('Energy (MeV)')
		ax.set_ylabel(isotope(istp).TeX()+' EOB Activity (kBq)')
		ax.legend(loc=0)
		f.tight_layout()
		if saveplot:
			f.savefig('../plots/monitors/'+istp+'_Al_Correction.png')
			f.savefig('../plots/monitors/'+istp+'_Al_Correction.pdf')
			plt.close()
		else:
			plt.show()
	def calc_chi2(self,table='mcnp'):
		if self.monitor_channels is None:
			self.calculate_current(False,table)
		mnch = [i for i in self.monitor_channels if i['foil']=='Cu10']
		monitor_dat = {str(i[0]):[float(i[1]),float(i[2]),float(i[3]),float(i[4])] for i in self.db.execute('SELECT * FROM monitor_data WHERE foil=?',('Cu10',))}
		return sum([(ch['I']-monitor_dat[ch['foil']][2])**2/ch['sigI']**2 for ch in mnch])
		# monitor_dat = {str(i[0]):[float(i[1]),float(i[2]),float(i[3]),float(i[4])] for i in self.db.execute('SELECT * FROM monitor_data')}
		# return sum([(ch['I']-monitor_dat[ch['foil']][2])**2/ch['sigI']**2 for ch in self.monitor_channels])/float(len(self.monitor_channels)-3)
	def calculate_current(self,saveplot=True,table='mcnp'):
		irr_time = 5844.0
		self.db.execute('DELETE FROM monitor_data')
		self.monitor_channels = []
		for fl in self.foil_ids:
			density,thickness,unc_density = [(float(i[3]),float(i[4]),float(i[5])) for i in self.db.execute('SELECT * FROM foils WHERE sample=?',(fl,))][0]
			Z = [int(str(i[2]).split(',')[0]) for i in self.db.execute('SELECT * FROM compounds WHERE compound=?',(fl[:2],))][0]
			gm_per_mol = [float(str(i[1]).split(',')[13]) for i in self.db.execute('SELECT * FROM ziegler_dat WHERE z=?',(Z,))][0]
			E,dE = self.get_eng_dist(fl,table)
			if E is None:
				continue
			eng,hist = self.zg.get_flux_spectrum(fl,table)
			for itp in self.sample_map[fl[:2]]:
				A0,sigA0 = self.get_A0(fl,itp)
				if A0 is None or (itp=='22NA' and fl=='Al02'):
					continue
				sigma,ip = self.xslib.interpolate(itp,'monitor'),isotope(itp)
				self.monitor_channels.append({'foil':fl,'istp':itp,'E':E,'dE':dE})
				Ip = A0*1.602e-7/((sum([i*sigma(eng[n]) for n,i in enumerate(hist)])/sum(hist))*(0.1*density*thickness*6.022e-1/gm_per_mol)*(1.0-np.exp(-ip.decay_const()*irr_time)))
				self.monitor_channels[-1]['I'] = Ip
				self.monitor_channels[-1]['sigI'] = np.sqrt(sigA0**2*(Ip/A0)**2+unc_density**2*(Ip/density)**2)
		x,y,dx = [c['E'] for c in self.monitor_channels],[c['I'] for c in self.monitor_channels],[c['dE'] for c in self.monitor_channels]
		m,b = self.simple_regression(x,y)
		dy = [np.sqrt(c['sigI']**2+c['dE']**2*m**2) for c in self.monitor_channels]
		fit,unc = curve_fit(lambda E_p,dI,I0:[dI*i+I0 for i in E_p],x,y,p0=[abs(m),b],sigma=dy,bounds=([0,-np.inf],[np.inf,np.inf]))
		for fl in [str(i[1]) for i in self.db.execute('SELECT * FROM foils') if str(i[2]) in ['La','Al','Cu'] and not i[1].startswith('E')]:
			E,dE = self.get_eng_dist(fl,table)
			self.db.execute('INSERT INTO monitor_data VALUES(?,?,?,?,?)',(fl,E,dE,fit[0]*E+fit[1],np.sqrt(unc[0][0]*E**2+unc[1][1]+2.0*E*unc[0][1])))
		self.db_connection.commit()
		if saveplot:
			f,ax = plt.subplots()
			clr = {'62ZN':self.pallate['r'],'22NA':self.pallate['b'],'24NA':self.pallate['p'],'63ZN':self.pallate['aq'],'58CO':self.pallate['gy']}
			lbl = {'62ZN':'Cu(p,x)','22NA':'Al(p,x)','24NA':'Al(p,x)','63ZN':'Cu(p,x)','58CO':'Cu(p,x)'}
			channels = list(set([c['istp'] for c in self.monitor_channels]))
			Erange = np.arange(min(x)-1.0,max(x),0.1)
			for ch in channels:
				x,y,dx,dy = [c['E'] for c in self.monitor_channels if c['istp']==ch],[c['I'] for c in self.monitor_channels if c['istp']==ch],[c['dE'] for c in self.monitor_channels if c['istp']==ch],[c['sigI'] for c in self.monitor_channels if c['istp']==ch]
				ax.errorbar(x,y,xerr=dx,yerr=dy,ls='None',marker='o',color=clr[ch],label=lbl[ch]+isotope(ch).TeX())
			ax.plot(Erange,[fit[0]*i+fit[1] for i in Erange],color=self.pallate['k'],label='Linear Fit')
			ax.plot(Erange,[fit[0]*i+fit[1]+np.sqrt(unc[0][0]*i**2+unc[1][1]+2.0*i*unc[0][1]) for i in Erange],ls='--',color=self.pallate['k'],label=r'$\pm 1\sigma_{fit}$')
			ax.plot(Erange,[fit[0]*i+fit[1]-np.sqrt(unc[0][0]*i**2+unc[1][1]+2.0*i*unc[0][1]) for i in Erange],ls='--',color=self.pallate['k'])
			ax.set_ylim(-0.5,ax.get_ylim()[1])
			ax.set_xlabel('Proton Energy (MeV)',fontsize=18)
			ax.set_ylabel('Monitor Current (nA)',fontsize=18)
			ax.legend(loc=0,ncol=2)
			f.tight_layout()
			f.savefig('../plots/monitors/current_norm_'+table+'.png')
			f.savefig('../plots/monitors/current_norm_'+table+'.pdf')
			plt.close()
	def minimize(self,table='az'):
		chi2 = []
		E0,dE0 = 57.0,(0.4 if table=='az' else 0.2)
		if E0>58:
			dp = [1.0,1.1,1.15,1.175,1.2,1.225,1.25,1.3] if table=='az' else [1.0,1.1,1.2,1.3,1.325,1.35,1.375,1.4]
		else:
			dp = [0.9,0.95,0.97,0.975,0.98,0.985,0.99,0.995,1.0,1.005,1.01,1.015,1.025,1.05] if table=='az' else [1.0,1.05,1.1,1.125,1.14,1.145,1.15,1.155,1.16,1.175,1.2]
		for d in dp:
			print 'Solving for dp=',d
			if table=='az':
				self.zg.solve_dEdx_AZMC(E0=E0,dE0=dE0,dp=d)
			else:
				self.zg.generate_mcnp_input(E0=E0,dE0=dE0,N=5e4,dp=d)
				self.zg.run_mcnp()
			self.monitor_channels = None
			chi2.append(self.calc_chi2(table))
		mn_dp = dp[chi2.index(min(chi2))]
		if table=='az':
			self.zg.solve_dEdx_AZMC(E0=E0,dE0=dE0,N=25000,dp=mn_dp)
		else:
			self.zg.generate_mcnp_input(E0=E0,dE0=dE0,N=1e5,dp=mn_dp)
			self.zg.run_mcnp()
		self.zg.plot_dEdx_AZMC(sample='La',table=table)
		self.calculate_current(table=table)
		f,ax = plt.subplots()
		ax.plot([100.0*(d-1.0) for d in dp],chi2,ls='--',ms=6.0,marker='o',color=self.pallate['k'],label='Results of Minimization')
		ax.plot([100.0*(mn_dp-1.0)],[min(chi2)],ms=10,ls='None',marker='o',color=self.pallate['r'],label=r'Min. $\chi^2_{\nu}$='+str(round(min(chi2),2))+r' at $\Delta \rho$='+str(mn_dp))
		ax.set_xlabel('Change in Density (%)')
		ax.set_ylabel(r'$\chi^2_{\nu}$ (a.u.)')
		ax.legend(loc=0)
		f.tight_layout()
		f.savefig('../plots/monitors/minimize_'+table+'.png')
		f.savefig('../plots/monitors/minimize_'+table+'.pdf')
		plt.close()
	def fit_monitor_peaks(self):
		self.xslib.fit_peaks('monitor')
	def fit_initial_activities(self,saveplot=True,logscale=False):
		self.xslib.fit_initial_activities(group='monitor',saveplot=saveplot,logscale=logscale)


if __name__=="__main__":
	cb = calibration()
	# cb.fits()
	cb.energy_calibration()
	# cb.efficiency_calibration()
	# cb.resolution_calibration()

	mn = monitors()
	# mn.fit_monitor_peaks()
	# mn.fit_initial_activities()
	# mn.calculate_current()
	# mn.plot_Al_correction('22NA')
	# mn.plot_Al_correction('24NA')
	mn.minimize('az')
	# mn.minimize('mcnp')

	xsl = XS_library()
	# xsl.fit_peaks('target')
	# xsl.fit_initial_activities()
	# xsl.fit_parent_daughter_activities()
	# xsl.fit_134CE_activity()
	xsl.plot_cross_sections()
	xsl.plot_cross_sections(prediction=False)
	xsl.save_as_csv()
	xsl.save_as_xlsx()
	xsl.save_as_tex()

# zg = ziegler()
# zg.generate_mcnp_input(dp=1.15,E0=57.0,dE0=0.4)
# zg.run_mcnp()
# zg.solve_dEdx_AZMC(N=10000,dp=1.0,E0=57.0,dE0=0.4)
# zg.plot_dEdx_AZMC(table='mcnp',saveplot=False)
# zg.plot_stopping_power_Z(13)

# mn = monitors()
# mn.calculate_current(saveplot=False,table='mcnp')

# xsl = XS_library()
# xsl.ratios()
# xsl.plot_energy_index()

# sp = spectrum('J_1cm_Al01.Spe')
# sp = spectrum('C03_15cm_Cu05.Spe')
# sp = spectrum('K_1cm_La02.Spe')
# sp = spectrum('H_1cm_La01.Spe')
# sp = spectrum('C_5cm_La01.Spe')
# sp = spectrum('I_1cm_La01.Spe')
# sp = spectrum('A_29p5cm_La01.Spe')
# sp = spectrum('F03_5cm_La01.Spe')
# sp = spectrum('E_5cm_Cu10.Spe')
# sp = spectrum('BH170825_10cm_Mn54.Spe')
# sp = spectrum('BA170825_50cm_Eu152.Spe')
# sp = spectrum('AX170825_40cm_Ba133.Spe')
# sp= spectrum('AL170825_10cm_Cs137.Spe')
# sp.plot_fits(logscale=True,subpeak=True,save=False,printout=True)
